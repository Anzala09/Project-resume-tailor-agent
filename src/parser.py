"""
parser.py — Load and merge option2_job_links.xlsx with option2_jobs.json.

Join key: xlsx["#"] == json["id"]
Returns a list of merged job dicts ready for the pipeline.
"""

import json
from pathlib import Path

import openpyxl


def load_jobs(xlsx_path: str, json_path: str) -> list[dict]:
    """
    Merge the Excel job links file with the JSON job details file.

    Args:
        xlsx_path: Path to option2_job_links.xlsx
        json_path: Path to option2_jobs.json

    Returns:
        List of merged job dicts with keys:
            id, title, company, url, resume_path,
            description, requirements, nice_to_have
    """
    # Load JSON job details, indexed by id
    with open(json_path, "r", encoding="utf-8") as f:
        json_data = json.load(f)

    json_index = {job["id"]: job for job in json_data["jobs"]}

    # Load Excel and join
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    jobs = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))

        job_id = row_data.get("#")
        if job_id is None:
            continue

        json_job = json_index.get(job_id)
        if json_job is None:
            print(f"  Warning: No JSON entry found for job id={job_id}, skipping.")
            continue

        merged = {
            "id": job_id,
            "title": row_data.get("Job Title") or json_job.get("title"),
            "company": row_data.get("Company") or json_job.get("company"),
            "url": row_data.get("URL") or json_job.get("url"),
            "resume_path": row_data.get("Resume Path", ""),
            "description": json_job.get("description", ""),
            "requirements": json_job.get("requirements", []),
            "nice_to_have": json_job.get("nice_to_have", []),
        }
        jobs.append(merged)

    return jobs
